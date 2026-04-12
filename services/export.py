"""
Export Service
--------------
Generates PDF and Excel exports of grading results (FR-40).
Exported files are saved to the submission folder with
consistent naming conventions (FR-30).
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def export_to_excel(results_data, output_path):
    """
    Export grading results to an Excel file (FR-40).
    
    The Excel file contains:
    - Student info at the top
    - Per-criterion breakdown with scores, bands, and feedback
    - Overall score at the bottom
    
    Compatible with Microsoft Excel (NFR-16).
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Grading Results"
        
        student = results_data['student_info']
        results = results_data['grading_result']
        
        # --- Styles ---
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        subheader_font = Font(bold=True, size=10)
        subheader_fill = PatternFill(start_color="D5E8F0", end_color="D5E8F0", fill_type="solid")
        normal_font = Font(size=10)
        score_font = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        
        # --- Student Info ---
        ws.merge_cells('A1:F1')
        ws['A1'] = 'FYP Grading Results'
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = 'Student Name:'
        ws['A3'].font = subheader_font
        ws['B3'] = student.get('name', '')
        
        ws['A4'] = 'Student Number:'
        ws['A4'].font = subheader_font
        ws['B4'] = student.get('number', '')
        
        ws['A5'] = 'Report:'
        ws['A5'].font = subheader_font
        ws['B5'] = student.get('report_filename', '')
        
        ws['A6'] = 'Date Graded:'
        ws['A6'].font = subheader_font
        ws['B6'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        
        reviewed = results.get('marker_reviewed', False)
        ws['A7'] = 'Status:'
        ws['A7'].font = subheader_font
        ws['B7'] = 'Marker Reviewed' if reviewed else 'AI Generated (Unreviewed)'
        
        # --- Overall Score ---
        ws.merge_cells('D3:F3')
        ws['D3'] = 'Overall Score'
        ws['D3'].font = Font(bold=True, size=12)
        ws['D3'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('D4:F4')
        ws['D4'] = results.get('overall_score', 0)
        ws['D4'].number_format = '0.0"%"'
        ws['D4'].font = Font(bold=True, size=20)
        ws['D4'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('D5:F5')
        ws['D5'] = f"Grade Band: {results.get('overall_grade_band', 'N/A')}"
        ws['D5'].alignment = Alignment(horizontal='center')
        
        # --- Criteria Table Header ---
        row = 9
        headers = ['Criterion', 'Weight', 'Score', 'Grade Band', 'Feedback', 'Evidence']
        col_widths = [30, 10, 10, 12, 50, 30]
        
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            ws.column_dimensions[chr(64 + col_idx)].width = width
        
        # --- Criteria Rows ---
        row += 1
        for i, criterion in enumerate(results.get('criteria_results', [])):
            row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid") if i % 2 == 0 else PatternFill(fill_type=None)
            
            # Column 1: Criterion name (text)
            cell = ws.cell(row=row, column=1, value=criterion.get('criterion_name', ''))
            cell.font = normal_font
            cell.alignment = wrap_alignment
            cell.border = thin_border
            if row_fill.fill_type:
                cell.fill = row_fill
            
            # Column 2: Weighting (NUMBER, not string)
            # We store the raw number and format the cell to show %
            cell = ws.cell(row=row, column=2, value=criterion.get('weighting', 0))
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='center', vertical='top')
            cell.border = thin_border
            cell.number_format = '0.00"%"'  # Displays as "6.67%" but stores as 6.67
            if row_fill.fill_type:
                cell.fill = row_fill
            
            # Column 3: Score (NUMBER)
            cell = ws.cell(row=row, column=3, value=criterion.get('score', 0))
            cell.font = score_font
            cell.alignment = Alignment(horizontal='center', vertical='top')
            cell.border = thin_border
            cell.number_format = '0'  # Whole number, no decimals
            if row_fill.fill_type:
                cell.fill = row_fill
            
            # Column 4: Grade band (text — this stays as string since it's a range like "60-69")
            cell = ws.cell(row=row, column=4, value=criterion.get('grade_band', ''))
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='center', vertical='top')
            cell.border = thin_border
            if row_fill.fill_type:
                cell.fill = row_fill
            
            # Column 5: Feedback (text)
            cell = ws.cell(row=row, column=5, value=criterion.get('feedback', ''))
            cell.font = normal_font
            cell.alignment = wrap_alignment
            cell.border = thin_border
            if row_fill.fill_type:
                cell.fill = row_fill
            
            # Column 6: Evidence (text)
            cell = ws.cell(row=row, column=6, value=criterion.get('evidence_location', ''))
            cell.font = normal_font
            cell.alignment = wrap_alignment
            cell.border = thin_border
            if row_fill.fill_type:
                cell.fill = row_fill
            
            row += 1
        
        # --- Summary ---
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'Overall Summary'
        ws[f'A{row}'].font = Font(bold=True, size=11)
        
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = results.get('overall_summary', '')
        ws[f'A{row}'].alignment = wrap_alignment
        
        # --- Plagiarism Flags ---
        plagiarism = results.get('plagiarism_flags', 'None detected')
        if plagiarism and plagiarism != 'None detected':
            row += 2
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = 'Plagiarism & Referencing Flags'
            ws[f'A{row}'].font = Font(bold=True, size=11, color="C0392B")
            
            row += 1
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = plagiarism
        
        # Save
        wb.save(output_path)
        return {'success': True, 'path': output_path}
    
    except Exception as e:
        return {'error': f'Failed to export Excel: {str(e)}'}


def export_to_pdf(results_data, output_path):
    """
    Export grading results to a PDF file (FR-40).
    
    Uses ReportLab to generate a professional-looking PDF
    with student info, scores, feedback, and summary.
    """
    try:
        student = results_data['student_info']
        results = results_data['grading_result']
        
        # Create the PDF document
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )
        
        # Styles
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=10
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=12,
            spaceAfter=6,
            spaceBefore=12
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=9,
            leading=12
        )
        
        feedback_style = ParagraphStyle(
            'Feedback',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#444444')
        )
        
        # Build content
        content = []
        
        # Title
        content.append(Paragraph('FYP Grading Results', title_style))
        content.append(Spacer(1, 5*mm))
        
        # Student info table
        reviewed = results.get('marker_reviewed', False)
        status = 'Marker Reviewed' if reviewed else 'AI Generated (Unreviewed)'
        
        info_data = [
            ['Student Name:', student.get('name', ''), 'Overall Score:', f"{results.get('overall_score', 0)}%"],
            ['Student Number:', student.get('number', ''), 'Grade Band:', results.get('overall_grade_band', 'N/A')],
            ['Date:', datetime.now().strftime('%Y-%m-%d %H:%M'), 'Status:', status],
            ['Report:', student.get('report_filename', ''), '', '']
        ]
        info_table = Table(info_data, colWidths=[80, 140, 80, 100])
        info_table.setStyle(TableStyle([
            # Font for all cells
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            # Bold labels in column 0 (Student Name:, Student Number:, etc.)
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            # Bold labels in column 2 (Overall Score:, Grade Band:, Status:)
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            # Big bold overall score number
            ('FONTSIZE', (3, 0), (3, 0), 14),
            ('FONTNAME', (3, 0), (3, 0), 'Helvetica-Bold'),
            # Padding between rows
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            # Merge report filename across columns 1-3 on the last row
            # SPAN merges cells: (col_start, row) to (col_end, row)
            # Row 3 (0-indexed), columns 1 to 3
            ('SPAN', (1, 3), (3, 3)),
        ]))
        content.append(info_table)
        content.append(Spacer(1, 8*mm))
        
        # Criteria table
        content.append(Paragraph('Criterion Breakdown', heading_style))
        
        # Table header
        criteria_data = [['Criterion', 'Wt', 'Score', 'Band', 'Feedback']]
        
        for criterion in results.get('criteria_results', []):
            # Wrap long text in Paragraph for proper word wrapping
            name = Paragraph(criterion.get('criterion_name', ''), feedback_style)
            feedback = Paragraph(criterion.get('feedback', ''), feedback_style)
            
            criteria_data.append([
                name,
                f"{criterion.get('weighting', '')}%",
                f"{criterion.get('score', '')}%",
                criterion.get('grade_band', ''),
                feedback
            ])
        
        criteria_table = Table(criteria_data, colWidths=[90, 30, 35, 35, 210])
        criteria_table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGNMENT', (0, 0), (-1, 0), 'CENTER'),
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGNMENT', (1, 1), (3, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            # Borders
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            # Alternating row colours
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        # Add alternating row colours
        for i in range(1, len(criteria_data)):
            if i % 2 == 0:
                criteria_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8F9FA'))
                ]))
        
        content.append(criteria_table)
        content.append(Spacer(1, 8*mm))
        
        # Summary
        content.append(Paragraph('Overall Summary', heading_style))
        content.append(Paragraph(results.get('overall_summary', ''), normal_style))
        
        # Plagiarism flags
        plagiarism = results.get('plagiarism_flags', 'None detected')
        if plagiarism and plagiarism != 'None detected':
            content.append(Spacer(1, 5*mm))
            content.append(Paragraph('Plagiarism & Referencing Flags', heading_style))
            content.append(Paragraph(plagiarism, normal_style))
        
        # Build PDF
        doc.build(content)
        return {'success': True, 'path': output_path}
    
    except Exception as e:
        return {'error': f'Failed to export PDF: {str(e)}'}


def export_results(results_data, submission_folder):
    """
    Export results as both PDF and Excel (FR-40).
    
    Files are saved to the submission folder with consistent naming (FR-30).
    Returns paths to both files.
    """
    student = results_data['student_info']
    student_number = student.get('number', 'UNKNOWN')
    student_name = student.get('name', 'Unknown').replace(' ', '_')
    
    # Consistent naming convention (FR-30)
    base_name = f"{student_number}_{student_name}_results"
    
    excel_path = os.path.join(submission_folder, f"{base_name}.xlsx")
    pdf_path = os.path.join(submission_folder, f"{base_name}.pdf")
    
    excel_result = export_to_excel(results_data, excel_path)
    pdf_result = export_to_pdf(results_data, pdf_path)
    
    return {
        'excel': excel_result,
        'pdf': pdf_result
    }